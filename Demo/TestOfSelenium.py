import re
import time

from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from selenium import webdriver
from selenium.webdriver.common.by import By

from Configs.OtherTools import GetMd5

rootUrl = r'https://www.baidu.com/'  # 根地址
testUrl = r'http://www.mafengwo.cn/yj/11754/1-0-1.html' # 测试地址
savePathXlsx = r'./XlsxData/DataDemo.xlsx'
dataTitle = ("名称", "地址", "电话", "营业时间", "评论数量", "评论用户名", "评论时间", "评论内容")

def SaveDataXlsx(datalist):
    wb = load_workbook(savePathXlsx)  # 创建文档对象
    ws = wb["Sheet1"]  # 调用工作表
    colHead = dataTitle

    existLen = len(ws["A"])
    for i, item in enumerate(datalist):
        for j in range(0, len(colHead)):
            txt = ILLEGAL_CHARACTERS_RE.sub(r'', item[j])
            ws.cell(row=i + existLen + 1, column=j + 1, value=txt)

    wb.save(savePathXlsx)  # 设置保存路径并保存
    print("\nSave infomation completed!!!\n\n\n")

# 构造一个webdriver对象 使用谷歌浏览器 用百度作为例子 使用保存的句柄进行跳转
def sample1():
    browser = webdriver.Chrome()
    mainHandle = browser.current_window_handle
    browser.get(rootUrl)

    linkButton = browser.find_element(By.XPATH, r'//a[text()="hao123"]')
    print('linkButton.text=' + linkButton.text)
    linkButton.click()

    print('main title=' + browser.title)
    for handle in browser.window_handles:
        if handle == mainHandle:
            continue
        browser.switch_to.window(handle)
        print('now title=' + browser.title)

    browser.close()
    browser.switch_to.window(mainHandle)
    print('after close title=' + browser.title)

    # print(type(browser.page_source))

    # input('Press <Enter> to close')
    # browser.back()

# 页内跳转可以通过点击和back直接跳转句柄
def sample2():
    browser = webdriver.Chrome()
    browser.get(rootUrl)
    print('before title=' + browser.title)
    mainInput = browser.find_element(By.XPATH, r'//input[@id="kw"]')
    mainInput.send_keys('hanhan')
    mainButton = browser.find_element(By.XPATH, r'//input[@id="su"]')
    mainButton.click()
    time.sleep(6)
    print('skip title=' + browser.title)

    browser.back()
    print('back title=' + browser.title)

# 搜索不存在的东西
def sample3():
    browser = webdriver.Chrome()
    browser.get(rootUrl)
    try:
        linkButton = browser.find_element(By.XPATH, r'//a[text()="haohoahaoahohfdsoia"]') # 搜索失败会直接报错
        print(linkButton)
    except Exception:
        print('Not Exist')

# 翻页返回
def sample4():
    browser = webdriver.Chrome()
    browser.get(r'https://you.ctrip.com/place/NanchangCounty2429.html')
    browser.get(rootUrl)

    # button1 = browser.find_element(By.XPATH, r'//ul[@class="cf"]/li/a[text()="景点"]')
    # button1.click()
    time.sleep(2)

    time.sleep(2)

    time.sleep(10)

def getDataTest(browser):
    # 翻页循环
    while True:
        # 页面内列表循环
        try:
            linkButtons = browser.find_element(By.XPATH, r'//a[@class="title-link"]')
            # linkButtons = browser.find_element(By.XPATH, r'//div[@class="img"]')
            for itemButton in linkButtons:
                dataList = getFirstPage(browser, itemButton) # 返回一个数据列表
                SaveDataXlsx(dataList)
        except Exception:
            print('This page have not datalist')

        # 判断翻页
        try:
            turnPageButton = browser.find_element(By.XPATH, r'//a[text()="后一页"]')
            turnPageButton.click()
            # browser.find_element(By.XPATH, r'//span[@class="iconfont icon-btn_right"]')
        except Exception:
            print("Turn Page Complete")
            break # 捕捉到异常则为没有下一页了

# 进入列表中的网页收集信息（页内跳转）
def getFirstPage(browser, itemButton):
    itemButton.click() # 进入网页
    print('parsing page ' + browser.title)

    # 获取公用数据
    publicData = []
    publicHtml = browser.page_source

    findName = re.compile(r'<!-- react-text: 26 -->.*"(.*?)".*<!-- /react-text -->')
    findAddress = re.compile(r'<!-- react-text: 60 -->.*"(.*?)".*<!-- /react-text -->')
    findPhoneNumber = re.compile(r'<!-- react-text: 64 -->.*"(.*?)".*<!-- /react-text -->')
    findBusinessHours = re.compile(r'<!-- react-text: 67 -->.*"(.*?)".*<!-- /react-text -->')
    findCommentNum = re.compile(r'<!-- react-text: 229 -->.*"(.*?)".*<!-- /react-text -->')

    try: # 名称
        name = re.findall(findName, publicHtml)[0]
        publicData.append(name)
    except Exception:
        print("Name Error!!!")
        publicData.append("")

    try: # 地址
        address = re.findall(findAddress, publicHtml)[0]
        publicData.append(address)
    except Exception:
        print("Address Error!!!")
        publicData.append("")

    try: # 电话
        phoneNumber = re.findall(findPhoneNumber, publicHtml)[0]
        publicData.append(phoneNumber)
    except Exception:
        print("Phone Number Error!!!")
        publicData.append("")

    try: # 营业时间
        businessHours = re.findall(findBusinessHours, publicHtml)[0]
        publicData.append(businessHours)
    except Exception:
        print("Business Hours Error!!!")
        publicData.append("")

    try: # 评论数量
        commentNum = re.findall(findCommentNum, publicHtml)[0]
        publicData.append(commentNum)
    except Exception:
        print("Comment Num Error!!!")
        publicData.append("")


    # 获取评论数据列表
    commentData = []
    while True:
        try:
            commentList = browser.find_elements(By.XPATH, r'//div[@class="com-cont"]/div/div[@class="list clear"]')
            for commentItem in commentList:
                itemList = []
                commentHtml = commentItem.text

                findCommentUsername = re.compile(r'<div class="name">(.*?)</div>')
                findCommentTime = re.compile(r'<div class="date">.*<span>(.*?)</span>.*</div>')
                findCommentContext = re.compile(r'<div class="desc">(.*)</div>', re.S)

                try: # 评论用户名
                    commentUsername = re.findall(findCommentUsername, commentHtml)[0]
                    itemList.append(commentUsername)
                except Exception:
                    print("Comment Username Error!!!")
                    publicData.append("")

                try: # 评论时间
                    commentTime = re.findall(findCommentTime, commentHtml)[0]
                    itemList.append(commentTime)
                except Exception:
                    print("Comment Time Error!!!")
                    publicData.append("")

                try: # 评论内容
                    commentContext = re.findall(findCommentContext, commentHtml)[0]
                    itemList.append(commentContext)
                except Exception:
                    print("Comment Context Error!!!")
                    publicData.append("")

                commentList.append(itemList) # 将一条评论的信息加入二维数组
        except Exception:
            print("No Comments")

        try:
            turnPageButton = browser.find_element(By.XPATH, r'//span[@class="iconfont icon-btn_right"]')
            turnPageButton.click()
        except:
            print("Comment Turn Page Complete")
            break

    # 拼接两种数据
    dataList = []
    for itemCommentData in commentData:
        dataList.append(publicData + itemCommentData) # 拼接

    browser.back() # 后退回到上一页
    return dataList

if __name__ == '__main__':
    # browser = webdriver.Chrome()
    # browser.get(testUrl)

    # sample1()
    # sample2()
    # sample3()
    sample4()
    # getDataTest(browser)