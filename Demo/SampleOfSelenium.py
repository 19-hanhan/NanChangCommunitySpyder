from lib2to3.pgen2 import driver
from threading import Thread
import time

from selenium import webdriver
from selenium.webdriver.support import expected_conditions as  EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup  # 解析网页源代码库
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from w3lib.html import remove_tags
import re  # 正则表达式库
import urllib.request, urllib.error  # 获取网页信息库
import xlwt  # xls文件操作库
import sqlite3  # db文件操作库

# 爬取数据信息参数
rootUrl = r'https://nc.meituan.com'  # 根地址
getUrl = r'/meishi/pn1/'  # 爬取列表位置根地址
usernameMeiTuan = r'13576090381'
passwordMeiTuan = r'lww000187'
savePathXlsx = r'./XlsxData/nc_meituan_com_food.xlsx'  # 保存地址
saveXlsxSheet = r'Sheet1' # 保存表
dataTitle = ("名称", "地址", "电话", "营业时间", "评论数量", "评论用户名", "评论时间", "评论内容")  # 数据字段名


# 操作函数
def main():
    ### 打开浏览器 ###
    browser = webdriver.Chrome()
    browser.get(rootUrl + getUrl)

    ### 登录账号 ###
    print('——————————Login Start——————————')
    LoginMeiTuan(browser, usernameMeiTuan, passwordMeiTuan)
    time.sleep(3)
    print('——————————Login Successful——————————')

    ### 爬取网页，获取数据列表 ###
    print('')
    print('——————————Crawling Start——————————')
    # GetData(browser)
    print('——————————Crawling Completed——————————')

    ### 关闭窗口 ###
    input('Press <Enter> to close')
    browser.close()

# 获取数据并保存（每收集了一家店保存一次）
def GetData(browser):
    # 翻页循环
    while True:
        # 页面内列表循环
        try:
            linkButtons = browser.find_element(By.XPATH, r'//div[@class="img"]')
            for itemButton in linkButtons:
                dataList = getFirstPage(browser, itemButton)  # 返回一个数据列表
                SaveDataXlsx(dataList)
        except Exception:
            print('This page have not datalist')

        # 判断翻页
        try:
            # turnPageButton = browser.find_element(By.XPATH, r'//a[text()="后一页"]')
            turnPageButton = browser.find_element(By.XPATH, r'//span[@class="iconfont icon-btn_right"]')
            turnPageButton.click()
            time.sleep(2)
        except Exception:
            break  # 捕捉到异常则为没有下一页了

# 跳转进入需要的第一级数据
def getFirstPage(browser, itemButton):
    itemButton.click() # 进入网页
    time.sleep(2)
    print('parsing page ' + browser.title)

    # 获取公用数据
    publicData = []
    publicHtml = browser.page_source

    findName = re.compile(r'<!-- react-text: 26 -->.*"(.*?)".*<!-- /react-text -->')
    findAddress = re.compile(r'<!-- react-text: 60 -->.*"(.*?)".*<!-- /react-text -->')
    findPhoneNumber = re.compile(r'<!-- react-text: 64 -->.*"(.*?)".*<!-- /react-text -->')
    findBusinessHours = re.compile(r'<!-- react-text: 67 -->.*"(.*?)".*<!-- /react-text -->')
    findCommentNum = re.compile(r'<!-- react-text: 229 -->.*"(.*?)".*<!-- /react-text -->')

    try: # 名称
        name = re.findall(findName, publicHtml)[0]
        publicData.append(name)
    except Exception:
        print("Name Error!!!")
        publicData.append("")

    try: # 地址
        address = re.findall(findAddress, publicHtml)[0]
        publicData.append(address)
    except Exception:
        print("Address Error!!!")
        publicData.append("")

    try: # 电话
        phoneNumber = re.findall(findPhoneNumber, publicHtml)[0]
        publicData.append(phoneNumber)
    except Exception:
        print("Phone Number Error!!!")
        publicData.append("")

    try: # 营业时间
        businessHours = re.findall(findBusinessHours, publicHtml)[0]
        publicData.append(businessHours)
    except Exception:
        print("Business Hours Error!!!")
        publicData.append("")

    try: # 评论数量
        commentNum = re.findall(findCommentNum, publicHtml)[0]
        publicData.append(commentNum)
    except Exception:
        print("Comment Num Error!!!")
        publicData.append("")


    # 获取评论数据列表
    commentData = []
    while True:
        try:
            commentList = browser.find_elements(By.XPATH, r'//div[@class="com-cont"]/div/div[@class="list clear"]')
            for commentItem in commentList:
                itemList = []
                commentHtml = commentItem.text

                findCommentUsername = re.compile(r'<div class="name">(.*?)</div>')
                findCommentTime = re.compile(r'<div class="date">.*<span>(.*?)</span>.*</div>')
                findCommentContext = re.compile(r'<div class="desc">(.*)</div>', re.S)

                try: # 评论用户名
                    commentUsername = re.findall(findCommentUsername, commentHtml)[0]
                    itemList.append(commentUsername)
                except Exception:
                    print("Comment Username Error!!!")
                    publicData.append("")

                try: # 评论时间
                    commentTime = re.findall(findCommentTime, commentHtml)[0]
                    itemList.append(commentTime)
                except Exception:
                    print("Comment Time Error!!!")
                    publicData.append("")

                try: # 评论内容
                    commentContext = re.findall(findCommentContext, commentHtml)[0]
                    itemList.append(commentContext)
                except Exception:
                    print("Comment Context Error!!!")
                    publicData.append("")

                commentList.append(itemList) # 将一条评论的信息加入二维数组
        except Exception:
            print("No Comments")

        try:
            turnPageButton = browser.find_element(By.XPATH, r'//span[@class="iconfont icon-btn_right"]')
            turnPageButton.click()
            time.sleep(2)
        except:
            print("Comment Turn Page Complete")
            break

    # 拼接两种数据
    dataList = []
    for itemCommentData in commentData:
        dataList.append(publicData + itemCommentData) # 拼接

    browser.back() # 后退回到上一页
    time.sleep(2)
    return dataList

# 传入浏览器对象、账号和密码登录美团
def LoginMeiTuan(brower, usr, pwd):
    print('登录信息录入中...')

    # 点击授权按钮
    agreeButton = brower.find_element(By.ID, 'user-agreement-wrap-text-circle')
    agreeButton.click()

    # # 输入账号
    # accountInput = brower.find_element(By.ID, 'login-email')
    # accountInput.send_keys(usr)
    #
    # # 输入密码
    # accountInput = brower.find_element(By.ID, 'login-password')
    # accountInput.send_keys(pwd)

    # 点击登录
    print('录入完成，正在登录...')
    # loginButton = brower.find_element(By.NAME, 'commit')
    loginButton = brower.find_element(By.ID, 'J-third-tencent')
    loginButton.click()
    time.sleep(60) # 手动扫码登录
    # try:
    #     print('清除弹窗中...')
    #     brower.switch_to.alert.accept()
    # except Exception:
    #     print('无弹窗')

# 保存数据添加到存在的xlsx文件中（全元素均为字符串类型）
def SaveDataXlsx(datalist):
    print("Save Data Start...")
    wb = load_workbook(savePathXlsx)  # 创建文档对象
    ws = wb[saveXlsxSheet]  # 调用工作表
    colHead = dataTitle

    existLen = len(ws["A"]) # 用第一列判定表中已经存在的行数
    for i, item in enumerate(datalist):
        for j in range(0, len(colHead)):
            txt = ILLEGAL_CHARACTERS_RE.sub(r'', item[j])
            ws.cell(row=i + existLen + 1, column=j + 1, value=txt)

    wb.save(savePathXlsx)  # 设置保存路径并保存
    print("Save Data Successful")

if __name__ == '__main__':
    main()
    print("\nWork done！")