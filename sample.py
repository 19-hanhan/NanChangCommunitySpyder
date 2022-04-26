import time

from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup # 解析网页源代码库
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from w3lib.html import remove_tags
import re # 正则表达式库
import urllib.request, urllib.error # 获取网页信息库
import xlwt # xls文件操作库
import sqlite3 # db文件操作库

rootUrl = "https://www.dianping.com"
getUrl = "/nanchang/ch10/p"  # 爬取的基础网址
# savePathXlsx = "./DataList.xlsx" # 保存的xlsx文件地址
# savePathDb = "./DataList.db" # 保存的数据库文件地址

# 操作函数
def main():
    # 爬取网页，获取数据列表
    # dataList = GetData(rootUrl + getUrl)

    # 保存数据
    # SaveDataXlsx(datalist, savePathXlsx)
    # SaveDataDb(datalist, savePathDb)
    pass

# 爬取网页每一页的数据
def GetData(url):
    dataList = []
    for page in range(1, 51):
        data = CheckFirst(url, page) # 返回值为一个二维列表
        dataList += data
    return dataList

# 爬取一级网页数据（获取标题与网页链接）
def CheckFirst(url, page):
    dataList = []
    html = AskUrl(url + str(page))  # 获取每个网页的html数据
    print("Enter page " + str(page))

    # 解析数据
    soup = BeautifulSoup(html, "html.parser")
    for pos, item in enumerate(soup.find_all("div", class_="shop-list J_shop-list shop-all-list")):  # 提取收集列表
        # print(item)
        firstData = [] # 一级页面所获得的数据（标题与链接的一维数组）
        item = str(item)  # 把网页源码转换成字符串

        # 创建正则表达式对象用于获取想要的信息
        findTitle = re.compile(r'<a.*data-click-name="shop_title_click".*><h4>(.*)</h4></a>')  # 查找标题名
        findLink = re.compile(r'<a.*data-click-name="shop_title_click".*href="(.*)">.*</a>')  # 创建网页链接正则表达式模式
        # findImgSrc = re.compile(r'<img.*src="(.*?)".*/>', re.S) # 让点符号可以匹配换行符，防止空链接
        # findTime = re.compile(r'<dd class="item-user">.*发表于.*([0-9][0-9][0-9][0-9]-[0-9][0-9]-[0-9][0-9]).*</dd>', re.S)
        # findView = re.compile(r'<i class="numview" title="浏览">(.*?)</i>')
        # findLove = re.compile(r'<i class="want" title="喜欢">(.*?)</i>')
        # findAnswer = re.compile(r'<i class="numreply" title="回复">(.*?)</i>')

        firstData.append(str(page) + '-' + str(pos + 1))  # 表示为第几页第几个数据

        # 通过创建的正则表达式对象获得信息
        try:
            title = re.findall(findTitle, item)[0]
            firstData.append(title)  # 添加标题
        except Exception:
            print("Title error!!!")
            firstData.append("")

        link = ''
        try:
            link = re.findall(findLink, item)[0]
            firstData.append(link)  # 添加网页链接
        except Exception:
            print("Link error!!!")
            firstData.append("")

        # 整合一级与二级返回数据
        if link == '':
            print("Getting page " + str(page) + " item " + str(pos + 1) + " unsuccessfully...[Error]") # 返回失败信息
            pass
        else:
            secondData = CheckSecond(link) # 二级网页查找内容（会返回一个二维列表）
            for item2 in secondData: # 取出一行的评论数据
                tmp = firstData + item2 # 评论数据加上该评论的用户数据信息
                dataList.append(tmp) # 将该行加入到返回列表中
            print("Getting page " + str(page) + " item " + str(pos + 1) + " successfully...") # 返回成功信息

    time.sleep(1) # 爬完一家店之后休息一秒
    return dataList

# 查找二级网页内容
def CheckSecond(url):
    html = AskUrl(url)

    # 解析数据
    soup = BeautifulSoup(html, "html.parser")
    for item in soup.find_all("h2", class_="mod-title J-tab"):  # 提取所有符合要求的模块的代码
        item = str(item)

        findLink = re.compile(r'<a class="item all-review" data-type="all" href="(.*)">.*</a>') # 查找三级链接

        link = ''
        try:
            link = re.findall(findLink, item)[0]
        except Exception:
            print("Link error!!!")

        if link == '':
            return []
        else:
            return CheckThird(rootUrl + link)

# 查找三级网页内容
def CheckThird(link):
    bd = ""
    html = AskUrl(link)

    # 解析数据
    soup = BeautifulSoup(html, "html.parser")
    for item in soup.find_all("div", class_="ctd_content"):  # 提取所有符合要求的模块的代码
        item = str(item)

        findText = re.compile(r'<p>(.*?)</p>', re.S) # 查找内容

        text = re.findall(findText, item)

        for tmpText in text:
            bd += tmpText

    pre = re.compile(r'<(.*?)>', re.S)
    bd = re.sub(pre, " ", bd)  # 去掉<>
    # print("Getting second successfully...")
    return bd

# 访问并获得网页源码
def AskUrl(url):
    head = { # 用于伪装的信息
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.114 Safari/537.36"
    }

    req = urllib.request.Request(url = url, headers = head) # 伪装后的申请信息
    html = "" # 网页源代码
    try:
        response = urllib.request.urlopen(req)
        html = response.read().decode("utf-8") # 用utf-8提取源码
        # print(html)
    except urllib.error.URLError as e:
        if hasattr(e, "code"):
            print(e.code)
        if hasattr(e, "reason"):
            print(e.reason)

    return html

# 保存数据到xlsx文件
def SaveDataXlsx(datalist, savePathXlsx):
    wb = load_workbook(savePathXlsx)  # 创建文档对象
    ws = wb["sheet1"]  # 调用工作表
    colHead = ("链接", "标题", "时间", "浏览量", "喜欢", "回复数", "内容")

    lsA = ws["A"]
    for i, item in enumerate(datalist):
        print("Saving item " + str(i + 1) + " successfully...")
        for j in range(0, len(colHead)):
            txt = ILLEGAL_CHARACTERS_RE.sub(r"", item[j])
            ws.cell(row = i + len(lsA) + 1, column = j + 1, value = txt)

    wb.save(savePathXlsx)  # 设置保存路径并保存
    print("\nSave infomation completed!!!\n\n\n")

if __name__ == '__main__':
    main()
    print("\nWork done！")
